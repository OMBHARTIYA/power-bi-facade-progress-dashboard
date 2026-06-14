import argparse
import re
import shutil
from pathlib import Path
from openpyxl import load_workbook


def load_project_ids(excel_path: Path, sheet_name: str, column_name: str) -> list[str]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        raise ValueError(f"Sheet '{sheet_name}' is empty")

    headers = ["" if h is None else str(h) for h in headers]
    if column_name not in headers:
        raise ValueError(f"Column '{column_name}' not found. Available columns: {headers}")

    idx = headers.index(column_name)
    seen = set()
    ids: list[str] = []
    for row in rows:
        if row is None or idx >= len(row):
            continue
        value = row[idx]
        if value is None:
            continue
        value = str(value).strip()
        if not value or value in seen:
            continue
        seen.add(value)
        ids.append(value)

    if not ids:
        raise ValueError(f"No project IDs found in column '{column_name}'")
    return ids


def choose_current_value(existing_text: str, ids: list[str], parameter_name: str) -> str:
    expr_pattern = rf'(expression\s+{re.escape(parameter_name)}\s*=\s*")([^"]+)(")'
    m = re.search(expr_pattern, existing_text)
    if m and m.group(2) in ids:
        return m.group(2)
    return ids[0]


def build_list_literal(ids: list[str]) -> str:
    escaped = [f'"{i.replace("\\", "\\\\").replace("\"", "\\\"")}"' for i in ids]
    return "{" + ", ".join(escaped) + "}"


def update_tmdl(text: str, parameter_name: str, ids: list[str], current_value: str) -> str:
    list_literal = build_list_literal(ids)

    expr_pattern = rf'(expression\s+{re.escape(parameter_name)}\s*=\s*")([^"]+)(")'
    text, expr_count = re.subn(expr_pattern, rf'\g<1>{current_value}\g<3>', text, count=1)
    if expr_count != 1:
        raise ValueError(f"Could not find expression for parameter '{parameter_name}'")

    list_pattern = r'List\s*=\s*\{.*?\}'
    text, list_count = re.subn(list_pattern, f'List={list_literal}', text, count=1, flags=re.DOTALL)
    if list_count != 1:
        raise ValueError("Could not find 'List={...}' block to update")

    return text


def main() -> None:
    parser = argparse.ArgumentParser(description="Update a ProjectID parameter list in a TMDL/TDML file from an Excel file.")
    parser.add_argument("--excel", required=True, help="Path to projects Excel file")
    parser.add_argument("--tmdl", required=True, help="Path to TMDL/TDML text file containing the ProjectID parameter")
    parser.add_argument("--sheet", default="projects", help="Excel sheet name (default: projects)")
    parser.add_argument("--column", default="items.id", help="Excel column name containing project IDs (default: items.id)")
    parser.add_argument("--parameter", default="ProjectID", help="Parameter name in TMDL/TDML (default: ProjectID)")
    parser.add_argument("--current-value", default=None, help="Optional current/default parameter value. If omitted, keep the existing value when possible, otherwise use the first ID.")
    parser.add_argument("--no-backup", action="store_true", help="Do not create a .bak backup file")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    tmdl_path = Path(args.tmdl)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if not tmdl_path.exists():
        raise FileNotFoundError(f"TMDL/TDML file not found: {tmdl_path}")

    ids = load_project_ids(excel_path, args.sheet, args.column)
    original_text = tmdl_path.read_text(encoding="utf-8")
    current_value = args.current_value or choose_current_value(original_text, ids, args.parameter)
    if current_value not in ids:
        raise ValueError(f"Current value '{current_value}' is not present in the Excel ID list")

    updated_text = update_tmdl(original_text, args.parameter, ids, current_value)

    if not args.no_backup:
        backup_path = tmdl_path.with_suffix(tmdl_path.suffix + ".bak")
        shutil.copy2(tmdl_path, backup_path)
        print(f"Backup created: {backup_path}")

    tmdl_path.write_text(updated_text, encoding="utf-8")
    print(f"Updated {args.parameter} with {len(ids)} IDs")
    print(f"Current value: {current_value}")
    print(f"File updated: {tmdl_path}")


if __name__ == "__main__":
    main()
